import pandas as pd
import openpyxl

class readdata():

    @staticmethod
    def getjumlahbaris(filepath:str,sheetname:str)->int:
        df = pd.read_excel(io=filepath,sheet_name=sheetname)
        jumlahbaris = df.loc[0,"jumlahbaris"]
        return jumlahbaris
    
    def bacanomormeter(filepath:str,sheetname:str,baris:int)->int:
        df = pd.read_excel(io=filepath,sheet_name=sheetname)
        nomormeter = df.loc[baris,"nomormeter"]
        return nomormeter
    
    def write_idpel(filepath:str,sheetname:str,data:str,column_number:int,row_number:int):
        workbook = openpyxl.load_workbook(filename=filepath)
        worksheet = workbook.get_sheet_by_name(sheetname)
        worksheet.cell(row=row_number,column=column_number).value = int(data)
        workbook.save(filename=filepath)
    def write_nomorgardu(filepath:str,sheetname:str,data:str,column_number:int,row_number:int):
        workbook = openpyxl.load_workbook(filename=filepath)
        worksheet = workbook.get_sheet_by_name(sheetname)
        worksheet.cell(row=row_number,column=column_number).value = data
        workbook.save(filename=filepath)
